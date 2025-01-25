"""
Data type conversion of different files
"""
import csv
import json
from itertools import islice
import codecs

import yaml
from openpyxl import load_workbook


def check_data(list_data: list) -> list:
    """
    Checking test data format.
    :param list_data:
    :return:
    """
    if isinstance(list_data, list) is False:
        raise TypeError("The data format is not `list`.")
    if len(list_data) == 0:
        raise ValueError("The data format cannot be `[]`.")
    if isinstance(list_data[0], dict):
        test_data = []
        for data in list_data:
            line = []
            for d in data.values():
                line.append(d)
            test_data.append(line)
        return test_data

    return list_data


def csv_to_list(file: str = None, line: int = 1, end_line: int = None) -> list:
    """
    Convert CSV file data to list
    :param file: Path to file
    :param line: Start line of read data
    :param end_line: End line of read data
    :return: list data

    Usage:
        csv_to_list("data.csv", line=1)
    """
    if file is None:
        raise FileExistsError("Please specify the CSV file to convert.")

    table_data = []
    with codecs.open(file, 'r', encoding='utf_8_sig') as csv_file:
        csv_data = csv.reader(csv_file)
        for i in islice(csv_data, line - 1, end_line):
            table_data.append(i)

    return table_data


def excel_to_list(file: str = None, sheet: str = "Sheet1", line: int = 1, end_line: int = None) -> list:
    """
    Convert Excel file data to list
    :param file: Path to file
    :param sheet: Excel sheet, default name is Sheet1
    :param line: Start line of read data
    :param end_line: Start line of read data
    :return: list data

    Usage:
        excel_to_list("data.xlsx", sheet="Sheet1", line=1)
    """
    if file is None:
        raise FileExistsError("Please specify the Excel file to convert.")

    excel_table = load_workbook(file)
    sheet = excel_table[sheet]
    if end_line is None:
        end_line = sheet.max_row

    table_data = []
    for i in sheet.iter_rows(line, end_line):
        line_data = []
        for field in i:
            line_data.append(field.value)
        table_data.append(line_data)

    return table_data


def json_to_list(file: str = None, key: str = None) -> list:
    """
    Convert JSON file data to list
    :param file: Path to file
    :param key: Specifies the key for the dictionary
    :return: list data

    Usage:
        json_to_list("data.yaml", key="login")
    """
    if file is None:
        raise FileExistsError("Please specify the JSON file to convert.")

    if key is None:
        with open(file, "r", encoding="utf-8") as json_file:
            data = json.load(json_file)
            list_data = check_data(data)
    else:
        with open(file, "r", encoding="utf-8") as json_file:
            try:
                data = json.load(json_file)[key]
                list_data = check_data(data)
            except KeyError as exc:
                raise ValueError(f"Check the test data, no '{key}'.") from exc

    return list_data


def yaml_to_list(file: str = None, key: str = None) -> list:
    """
    Convert YAML file data to list
    :param file: Path to file
    :param key: Specifies the key for the dictionary
    :return: list data

    Usage:
        yaml_to_list("data.yaml", key="login")
    """
    if file is None:
        raise FileExistsError("Please specify the YAML file to convert.")

    if key is None:
        with open(file, "r", encoding="utf-8") as yaml_file:
            data = yaml.load(yaml_file, Loader=yaml.FullLoader)
            list_data = check_data(data)
    else:
        with open(file, "r", encoding="utf-8") as yaml_file:
            try:
                data = yaml.load(yaml_file, Loader=yaml.FullLoader)[key]
                list_data = check_data(data)
            except KeyError as exc:
                raise ValueError(f"Check the YAML test data, no '{key}'") from exc

    return list_data
